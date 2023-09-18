import constants as keys
import response as R
import openpyxl
from datetime import datetime
from telegram import Update
from telegram.ext import Updater, CommandHandler, MessageHandler, CallbackContext, Filters, ConversationHandler

total_salary = 0
# lưu dữ liệu vào file excel
def import_data(infor1, infor2, col1, col2):
    cur_date = datetime.now().day
    
    # Ghi thông tin vào tệp Excel
    wb = openpyxl.load_workbook('payroll.xlsx')
    sheet = wb.active

    sheet.cell(row = cur_date+2, column = col1).value = infor1
    sheet.cell(row = cur_date+2, column = col2).value = infor2

    global total_salary
    if sheet['B34'].data_type == 'f':
        total_salary = sheet.cell(row = 34, column = 2).value
    
    wb.save('payroll.xlsx')
    wb.close()

# Hàm xử lý khi người dùng gửi thông tin lương
def handle_salary_info(update: Update, context: CallbackContext):
    global total_salary
    # Trích xuất thông tin từ tin nhắn của người dùng
    message_text = update.message.text
    
    # Tách thông tin lớp và sĩ số bằng dấu chấm phẩy và dấu cách
    try:
        data = message_text.split()
        if len(data) == 3:
            type = data[0]
            infor1 = data[1]
            infor2 = data[2]
            if type == "tg":
                import_data(infor1, infor2, 2, 3)   # trợ giảng
            else:
                import_data(infor1, infor2, 5, 6)   # phụ đạo
            update.message.reply_text(f'Tổng lúa đã lụm {total_salary}')
        
        elif len(data) == 2:
            type = data[0]
            infor1 = int(data[1])
            import_data(infor1, infor1, 8, 8)       # gia sư
            update.message.reply_text(f'Tổng lúa đã lụm {total_salary}')
        else:
            update.message.reply_text('Gửi đúng định dạng đê!')
    except Exception:
        update.message.reply_text('Tao không đọc được m viết gì!')
        
# Hàm khởi động với các câu hỏi trong file Response.py
def handle_message(update: Update, context: CallbackContext):
    text = str(update.message.text).lower()
    response = R.sample_response(text)
    update.message.reply_text(response)
  
  
def main():
    updater = Updater(keys.API_KEY, use_context=True)
    dp = updater.dispatcher
    dp.add_handler(MessageHandler(Filters.text, handle_salary_info))

    updater.start_polling()
    updater.idle()


if __name__ == '__main__':
    main()





# # Hàm khởi động với lệnh /start trong telegram
# def start_command(update: Update, context: CallbackContext) -> None:
#     update.message.reply_text(f'Hello {update.effective_user.first_name}')

# # Hàm khởi động với lệnh /help trong telegram
# def help_command(update: Update, context: CallbackContext) -> None:
#     update.message.reply_text("Bạn muốn tôi giúp gì?")
    

# # Param thứ nhất là nội dung câu lệnh, param thứ 2 là hàm để chạy lệnh đó
    # dp.add_handler(CommandHandler('start', start_command))
    # dp.add_handler(CommandHandler("help", help_command))
    # dp.add_handler(MessageHandler(Filters.text, handle_message))